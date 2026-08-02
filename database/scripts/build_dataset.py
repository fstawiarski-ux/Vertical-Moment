#!/usr/bin/env python3
"""Vertical Moment — routes dataset builder.

Single source of truth: vertical_moment_master_routes.xlsx -> sheet `Routes`.
Generates data/ output: routes.json, routes.csv, facets.json, areas.json,
by-area/<slug>.json, stats.json, index.json. Idempotent; overwrites data/ only.

Usage:
    python3 build_dataset.py [master.xlsx] [out_dir]   # defaults: master in cwd, out=data/
"""
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

COLUMNS = [
    "Route", "Grade", "Grade band", "Discipline", "Grade system", "Area",
    "Wall", "Source", "Latitude", "Longitude", "Row Key", "Notion Page ID",
    "Status", "Coord Source",
]


def slugify(name: str) -> str:
    """Fold umlauts (oe/ae/ue) and strip unsafe chars for filenames/URLs."""
    s = name or ""
    s = s.replace("\u00e4", "ae").replace("\u00f6", "oe").replace("\u00fc", "ue")
    s = s.replace("\u00c4", "Ae").replace("\u00d6", "Oe").replace("\u00dc", "Ue")
    s = s.replace("\u00df", "ss")
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r"[^A-Za-z0-9-_]+", "-", s).strip("-").lower()
    return s or "unnamed"


def load_master(path: Path) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Routes" not in wb.sheetnames:
        raise SystemExit(f"No 'Routes' sheet in {path}")
    ws = wb["Routes"]
    header = [c.value for c in ws[1]]
    if header[:13] != COLUMNS[:13]:
        raise SystemExit(f"Unexpected header in {path}: {header}")
    routes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[6]:
            continue
        rec = {
            "route": row[0],
            "grade": row[1],
            "grade_band": row[2],
            "discipline": row[3],
            "grade_system": row[4],
            "area": row[5],
            "wall": row[6],
            "source": row[7],
            "latitude": row[8],
            "longitude": row[9],
            "row_key": row[10],
            "notion_page_id": row[11],
            "status": row[12],
        }
        # Column 14 is an optional additive provenance field (Coord Source).
        if len(header) > 13 and row[13]:
            rec["coord_source"] = row[13]
        routes.append(rec)
    return routes


def build(master: Path, out: Path) -> None:
    routes = load_master(master)
    out.mkdir(parents=True, exist_ok=True)

    # --- routes.json ------------------------------------------------------
    fields = [
        "route", "grade", "grade_band", "discipline", "grade_system", "area",
        "wall", "source", "latitude", "longitude", "row_key",
        "notion_page_id", "status",
    ]
    if any("coord_source" in r for r in routes):
        fields.append("coord_source")
    payload = {
        "meta": {
            "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00"),
            "count": len(routes),
            "fields": fields,
        },
        "routes": routes,
    }
    (out / "routes.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    # --- routes.csv -------------------------------------------------------
    import csv
    import io as _io
    csv_cols = fields
    buf = _io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(csv_cols)
    for r in routes:
        w.writerow("" if (v := r.get(c)) is None else str(v) for c in csv_cols)
    (out / "routes.csv").write_text(buf.getvalue(), encoding="utf-8")

    # --- facets.json ------------------------------------------------------
    def facet(key):
        c = Counter(r.get(key) or "" for r in routes)
        return {"options": {k: v for k, v in sorted(c.items())}, "total": len(routes)}

    facets = {
        "areas": facet("area"),
        "walls": facet("wall"),
        "grade_bands": facet("grade_band"),
        "disciplines": facet("discipline"),
        "grade_systems": facet("grade_system"),
        "sources": facet("source"),
        "statuses": facet("status"),
    }
    (out / "facets.json").write_text(
        json.dumps(facets, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    # --- areas.json -------------------------------------------------------
    areas = defaultdict(lambda: {"walls": defaultdict(lambda: Counter()), "routes": 0})
    for r in routes:
        a = areas[r["area"]]
        a["routes"] += 1
        a["walls"][r["wall"]][r["discipline"]] += 1
    area_tree = []
    for name, a in sorted(areas.items()):
        walls = []
        for wname, disp in sorted(a["walls"].items()):
            walls.append({
                "name": wname,
                "slug": slugify(wname),
                "count": sum(disp.values()),
                "disciplines": dict(sorted(disp.items())),
            })
        area_tree.append({
            "name": name,
            "slug": slugify(name),
            "count": a["routes"],
            "walls": walls,
        })
    (out / "areas.json").write_text(
        json.dumps(area_tree, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    # --- by-area/<slug>.json ----------------------------------------------
    by_area_dir = out / "by-area"
    by_area_dir.mkdir(parents=True, exist_ok=True)
    per_area = defaultdict(list)
    for r in routes:
        per_area[r["area"]].append(r)
    for name, rs in per_area.items():
        (by_area_dir / f"{slugify(name)}.json").write_text(
            json.dumps({"area": name, "count": len(rs), "routes": rs},
                       ensure_ascii=False, indent=1),
            encoding="utf-8",
        )

    # --- stats.json --------------------------------------------------------
    stats = {
        "total_routes": len(routes),
        "total_areas": len(areas),
        "total_walls": len({r["wall"] for r in routes}),
        "with_coordinates": sum(1 for r in routes if r.get("latitude") is not None),
        "by_status": dict(Counter(r.get("status") or "" for r in routes)),
        "by_discipline": dict(Counter(r.get("discipline") or "" for r in routes)),
        "by_grade_system": dict(Counter(r.get("grade_system") or "" for r in routes)),
        "per_area": {n: a["routes"] for n, a in sorted(areas.items())},
    }
    (out / "stats.json").write_text(
        json.dumps(stats, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    # --- index.json --------------------------------------------------------
    index = {
        "generated": payload["meta"]["generated"],
        "files": [
            {"file": "routes.json", "kind": "routes", "count": len(routes)},
            {"file": "routes.csv", "kind": "csv", "count": len(routes)},
            {"file": "facets.json", "kind": "facets", "count": len(facets)},
            {"file": "areas.json", "kind": "areas", "count": len(area_tree)},
            {"file": "stats.json", "kind": "stats", "count": 1},
        ]
        + [
            {"file": f"by-area/{slugify(n)}.json", "kind": "by-area",
             "area": n, "count": len(rs)}
            for n, rs in sorted(per_area.items())
        ],
        "areas": [a["slug"] for a in area_tree],
    }
    (out / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    print(f"built {len(routes)} routes -> {out}")
    print("  files: routes.json, routes.csv, facets.json, areas.json, "
          "by-area/*.json, stats.json, index.json")


if __name__ == "__main__":
    args = sys.argv[1:]
    cwd = Path.cwd()
    master = Path(args[0]) if args else cwd / "vertical_moment_master_routes.xlsx"
    out = Path(args[1]) if len(args) > 1 else cwd / "data"
    build(master, out)
